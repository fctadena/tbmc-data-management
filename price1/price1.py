import pandas as pd
import time
import os
from datetime import datetime
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests
import re

#2. DEVELOP THE APP
#2.1 create necessary variables, flags & dataframe for scrapped data "scrapped_data_df"
file_path = "TBMC_ITEMS_TO_MONITOR.xlsx"
# columns = ['description', 'type', 'url', 'price', 'date_stamp']
# scrapped_data_df = pd.DataFrame(columns=columns)



#2.2 xlsx to dict - initialize the process by storing the required items (products to monitor) in dictionary called "products"
def read_excel_file(file_path):
    try:
        df = pd.read_excel(file_path)
    except FileNotFoundError:
        print("File not found.")
        return {}
    
    expected_headers = ["description", "type", "url", "classs"]
    actual_headers = df.columns.tolist()
    
    if actual_headers != expected_headers:
        print("Headers are not as expected.")
        return {}
    
    products = []
    for index, row in df.iterrows():
        description = row["description"]
        product_type = row["type"]
        url = row["url"]
        classs = row["classs"]
        
        product = {
            "description": description,
            "type": product_type,
            "url": url,
            "classs": classs
        }
        
        products.append(product)
    
    return products



# #2.3 main_processor - loop through the dict, perform scrapping, and add records to scrapped_data_df
def core_func(products):
    # columns = ["description", "type", "url", "price", "time_stamp", "status"]
    columns = ["description", "type", "url", "classs", "price", "time_stamp"]
    scrapped_data_df = pd.DataFrame(columns=columns)

    
    
    for i in products:
        #Generate "price", "time_stamp" and "status" here.!!!
        time_stamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        try:
            price_source = requests.get(i['url'])
            soup_price_source = BeautifulSoup(price_source.text, 'html')
            price = float(re.sub(r'[^\d.]+', '', soup_price_source.find('span', attrs={'class':i['classs']}).text.strip()))
        except:
            print("Exept on try")
            price = 0
            
        
        new_row = [
            i['description'],
            i['type'],
            i['url'],
            i['classs'],
            price,
            time_stamp
        ]
        
        print("test")
        print(new_row)

        scrapped_data_df.loc[len(scrapped_data_df)] = new_row
        
    print(scrapped_data_df.dtypes)
    return scrapped_data_df



#2.4 append (or create PRICE_RECORDINGS.xlsx) using the dataframe
def update_price_recordings(scrapped_data_df):
    file_name = "PRICE_RECORDINGS_NEW1.xlsx"
    sheet_name = 'Sheet1'
    
    # Check if the file exists
    if not os.path.exists(file_name):
        # If the file does not exist, create it and write the scrapped_data_df to it
        try:
            with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
                scrapped_data_df.to_excel(writer, index=False, sheet_name=sheet_name)
        except Exception as e:
            print(f"Error creating file: {e}")
            return False
    else:
        try:
            # If the file exists, load it and append the data
            book = load_workbook(file_name)
            
            # Ensure at least one sheet is visible
            if all(sheet.sheet_state == 'hidden' for sheet in book.worksheets):
                book.create_sheet(title=sheet_name)
            
            with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                
                # Check if the specified sheet exists
                if sheet_name in writer.sheets:
                    # Get the last row in the existing Excel sheet
                    startrow = writer.sheets[sheet_name].max_row
                else:
                    startrow = 0
                
                # Append the data
                scrapped_data_df.to_excel(writer, startrow=startrow, index=False, header=startrow==0, sheet_name=sheet_name)
                
                # Save and close the file
                writer.save()
        except Exception as e:
            print(f"Error appending data: {e}")
            return False
    
    return True


#2.5 create main function 
def main():
    # # 2.1 create necessary variables, flags & dataframe for scrapped data "scrapped_data_df"
    # file_path = "TBMC_ITEMS_TO_MONITOR.xlsx"
    # columns = ['description', 'type', 'url', 'price', 'date_stamp']
 

    # scrapped_data_df = pd.DataFrame(columns=columns)

    # 2.2 xlsx to dict - initialize the process by storing the required items (products to monitor) in dictionary called "products"
    products = read_excel_file(file_path)

    # 2.3 main_processor - loop through the dict, perform scrapping, and add records to scrapped_data_df
    scrapped_data_df = core_func(products)

    # 2.4 append (or create PRICE_RECORDINGS.xlsx) using the dataframe
    update_price_recordings(scrapped_data_df)

if __name__ == "__main__":
    main()